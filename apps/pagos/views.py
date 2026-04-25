from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Count, Avg
from django.db import transaction, IntegrityError
from django.http import HttpResponse, JsonResponse
from decimal import Decimal
import datetime, json

from .models import CajaDiaria, MovimientoCaja, AjusteStock, InsumoStock, Pago, Caja, CajaSesion, MovimientoCajaSesion
from apps.pedidos.models import Pedido, ItemPedido, ItemPedidoSabor
from apps.productos.models import Sabor, Producto
from django.views.decorators.http import require_POST, require_GET


# ─────────────────────────────────────────────
# CAJA
# ─────────────────────────────────────────────

def caja_hoy(request):
    hoy  = timezone.now().date()
    caja = CajaDiaria.objects.filter(fecha=hoy).first()

    if caja:
        movimientos = caja.movimientos.all()
        ingresos = caja.movimientos.filter(
            tipo=MovimientoCaja.TIPO_INGRESO
        ).aggregate(total=Sum("monto"))["total"] or 0

        egresos = caja.movimientos.filter(
            tipo=MovimientoCaja.TIPO_EGRESO
        ).aggregate(total=Sum("monto"))["total"] or 0
    else:
        movimientos = []
        ingresos = 0
        egresos = 0

    context = {
        "caja": caja,
        "hoy": hoy,
        "movimientos": movimientos,
        "ingresos": ingresos,
        "egresos": egresos,
    }
    return render(request, "pagos/caja.html", context)


def abrir_caja(request):
    if request.method == "POST":
        hoy = timezone.now().date()
        if CajaDiaria.objects.filter(fecha=hoy).exists():
            messages.warning(request, "Ya hay una caja abierta para hoy.")
            return redirect("pagos:caja_hoy")

        monto_inicial = Decimal(request.POST.get("monto_inicial", "0"))
        CajaDiaria.objects.create(fecha=hoy, monto_inicial=monto_inicial)
        messages.success(request, f"Caja abierta con ${monto_inicial} iniciales.")
    return redirect("pagos:caja_hoy")


def cerrar_caja(request):
    if request.method == "POST":
        hoy  = timezone.now().date()
        caja = get_object_or_404(CajaDiaria, fecha=hoy, cerrada=False)
        monto_real = Decimal(request.POST.get("monto_real", "0"))
        caja.cerrar(monto_real)
        messages.success(request, "Caja cerrada correctamente.")
    return redirect("pagos:caja_hoy")


def registrar_egreso(request):
    if request.method == "POST":
        hoy  = timezone.now().date()
        caja = CajaDiaria.objects.filter(fecha=hoy, cerrada=False).first()
        if not caja:
            messages.error(request, "No hay caja abierta hoy.")
            return redirect("pagos:caja_hoy")

        monto       = Decimal(request.POST.get("monto", "0"))
        descripcion = request.POST.get("descripcion", "Egreso manual")

        MovimientoCaja.objects.create(
            caja        = caja,
            tipo        = MovimientoCaja.TIPO_EGRESO,
            monto       = monto,
            descripcion = descripcion,
        )
        caja.calcular_cierre_esperado()
        caja.save(update_fields=["monto_cierre_esperado"])
        messages.success(request, f"Egreso de ${monto} registrado.")
    return redirect("pagos:caja_hoy")


# ─────────────────────────────────────────────
# MÉTRICAS
# ─────────────────────────────────────────────

def metricas(request):
    # Rango de fechas (default: últimos 30 días)
    hoy        = timezone.now().date()
    fecha_desde = request.GET.get("desde", str(hoy - datetime.timedelta(days=30)))
    fecha_hasta = request.GET.get("hasta", str(hoy))

    try:
        desde = datetime.date.fromisoformat(fecha_desde)
        hasta = datetime.date.fromisoformat(fecha_hasta)
    except ValueError:
        desde = hoy - datetime.timedelta(days=30)
        hasta = hoy

    pedidos_qs = Pedido.objects.filter(
        fecha_creacion__date__gte=desde,
        fecha_creacion__date__lte=hasta,
    ).exclude(estado=Pedido.ESTADO_CANCELADO)

    # KPIs principales
    total_ventas   = pedidos_qs.aggregate(Sum("total"))["total__sum"] or Decimal("0")
    cantidad_pedidos = pedidos_qs.count()
    ticket_promedio  = (total_ventas / cantidad_pedidos) if cantidad_pedidos else Decimal("0")

    # Ventas por método de pago
    ventas_por_pago = (
        pedidos_qs
        .values("metodo_pago_principal")
        .annotate(total=Sum("total"), cantidad=Count("id"))
        .order_by("-total")
    )

    # Productos más vendidos
    productos_top = (
        ItemPedido.objects
        .filter(pedido__in=pedidos_qs)
        .values("producto__nombre", "producto__tipo")
        .annotate(unidades=Sum("cantidad"), ingresos=Sum("subtotal"))
        .order_by("-unidades")[:10]
    )

    # Sabores más pedidos
    sabores_top = (
        ItemPedidoSabor.objects
        .filter(item_pedido__pedido__in=pedidos_qs)
        .values("sabor__nombre")
        .annotate(apariciones=Count("id"))
        .order_by("-apariciones")[:10]
    )

    # Pedidos por día (para gráfico)
    pedidos_por_dia = (
        pedidos_qs
        .values("fecha_creacion__date")
        .annotate(total_dia=Sum("total"), cantidad_dia=Count("id"))
        .order_by("fecha_creacion__date")
    )

    context = {
        "desde":            desde,
        "hasta":            hasta,
        "total_ventas":     total_ventas,
        "cantidad_pedidos": cantidad_pedidos,
        "ticket_promedio":  ticket_promedio,
        "ventas_por_pago":  ventas_por_pago,
        "productos_top":    productos_top,
        "sabores_top":      sabores_top,
        "pedidos_por_dia":  pedidos_por_dia,
    }
    return render(request, "pagos/metricas.html", context)


# ─────────────────────────────────────────────
# STOCK
# ─────────────────────────────────────────────

def stock(request):
    from apps.productos.models import Sabor
    from apps.pedidos.models import ItemPedidoSabor

    # Período de análisis para proyección: últimos 14 días
    hoy      = timezone.now().date()
    hace_14  = hoy - datetime.timedelta(days=14)

    # Consumo promedio diario por sabor (en kg) basado en ventas reales
    consumo_qs = (
        ItemPedidoSabor.objects
        .filter(item_pedido__pedido__fecha_creacion__date__gte=hace_14)
        .values("sabor_id")
        .annotate(apariciones=Count("id"))
    )
    consumo_map = {}
    for row in consumo_qs:
        # Aproximamos: cada aparición ≈ 0.05 kg (ajustable)
        kg_por_dia = (row["apariciones"] * 0.05) / 14
        consumo_map[row["sabor_id"]] = round(kg_por_dia, 4)

    sabores_data = []
    alertas      = []

    for sabor in Sabor.objects.all().order_by("nombre"):
        consumo_diario = consumo_map.get(sabor.id, 0)
        if consumo_diario > 0:
            dias_restantes = sabor.stock_kg / Decimal(str(consumo_diario))
            dias_restantes = int(dias_restantes)
        else:
            dias_restantes = None   # sin datos de venta

        alerta = not sabor.disponible or (dias_restantes is not None and dias_restantes <= 3)

        sabores_data.append({
            "sabor":           sabor,
            "consumo_diario":  consumo_diario,
            "dias_restantes":  dias_restantes,
            "alerta":          alerta,
        })

        if alerta:
            alertas.append(sabor)

    insumos  = InsumoStock.objects.all()
    insumos_alerta = [i for i in insumos if i.bajo_stock]

    context = {
        "sabores_data":    sabores_data,
        "insumos":         insumos,
        "alertas":         alertas,
        "insumos_alerta":  insumos_alerta,
        "hay_alertas":     bool(alertas or insumos_alerta),
    }
    return render(request, "pagos/stock.html", context)


def ajuste_stock(request):
    if request.method == "POST":
        sabor_id    = request.POST.get("sabor_id")
        cantidad_kg = Decimal(request.POST.get("cantidad_kg", "0"))
        motivo      = request.POST.get("motivo", "")

        sabor = get_object_or_404(Sabor, pk=sabor_id)
        ajuste = AjusteStock.objects.create(
            sabor       = sabor,
            cantidad_kg = cantidad_kg,
            motivo      = motivo,
        )
        ajuste.aplicar()
        messages.success(
            request,
            f"Stock de {sabor.nombre} ajustado. Nuevo stock: {sabor.stock_kg} kg"
        )
    return redirect("pagos:stock")


def alerta_stock_email(request):
    """Envía un email de alerta de stock crítico al dueño. Llamado por AJAX."""
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({"ok": False, "error": "Método no permitido"}, status=405)
 
    from apps.core.emails import notificar_stock_critico
    from django.http import JsonResponse
 
    sabores_criticos = []
    for sabor in Sabor.objects.filter(activo=True):
        if not sabor.disponible:
            sabores_criticos.append({"sabor": sabor, "motivo": "agotado"})
 
    insumos_criticos = [i for i in InsumoStock.objects.all() if i.bajo_stock]
 
    try:
        notificar_stock_critico(sabores_criticos, insumos_criticos)
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})
 

# ─────────────────────────────────────────────
# EXPORTAR EXCEL PARA PROVEEDORES
# ─────────────────────────────────────────────

def exportar_proveedores(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl no instalado. Ejecutá: pip install openpyxl", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedido a proveedores"

    # Estilos
    header_fill = PatternFill("solid", fgColor="3D1C02")
    header_font = Font(color="FFF8F0", bold=True)
    turquesa_fill = PatternFill("solid", fgColor="2EC4B6")

    headers = ["Sabor", "Stock actual (kg)", "Stock mínimo (kg)", "Sugerido pedir (kg)", "Cantidad a pedir"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sabores = Sabor.objects.filter(activo=True).order_by("nombre")
    for row, sabor in enumerate(sabores, 2):
        # Sugerimos pedir el doble del mínimo menos lo que ya hay
        sugerido = max(Decimal("0"), (sabor.stock_minimo_kg * 2) - sabor.stock_kg)

        ws.cell(row=row, column=1, value=sabor.nombre)
        ws.cell(row=row, column=2, value=float(sabor.stock_kg))
        ws.cell(row=row, column=3, value=float(sabor.stock_minimo_kg))
        ws.cell(row=row, column=4, value=float(sugerido))
        # Columna editable para que el dueño ponga la cantidad real
        ws.cell(row=row, column=5, value=float(sugerido)).fill = turquesa_fill

    # Ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pedido_proveedores.xlsx"'
    wb.save(response)
    return response

# ─────────────────────────────────────────────
# VISTAS AJAX PARA POS
#────────────────────────────────────────────
@require_GET
def pos_movimientos(request):
    import datetime
    from apps.pagos.models import CajaSesion, MovimientoCajaSesion, Caja, Pago

    sesion_id = request.GET.get("sesion_id")
    fecha_str = request.GET.get("fecha", str(timezone.now().date()))

    try:
        fecha = datetime.date.fromisoformat(fecha_str)
    except ValueError:
        fecha = timezone.now().date()

    # ── Determinar qué sesiones mostrar ──────────────────────────
    if sesion_id:
        sesiones = CajaSesion.objects.filter(pk=sesion_id)
    else:
        # Todas las sesiones del día (puede haber más de una por turnos)
        sesiones = CajaSesion.objects.filter(
            fecha_apertura__date=fecha
        )

    movimientos  = []
    total_ventas  = Decimal("0")
    total_egresos = Decimal("0")

    for sesion in sesiones.order_by("fecha_apertura"):
        for mov in sesion.movimientos_sesion.select_related("pedido").order_by("fecha"):
            entry = {
                "hora":        mov.fecha.strftime("%H:%M"),
                "tipo":        mov.tipo,
                "total":       float(mov.monto),
                "descripcion": mov.descripcion,
                "numero":      mov.pedido.numero if mov.pedido else None,
                "metodo":      mov.pedido.metodo_pago_principal if mov.pedido else None,
                "origen":      mov.pedido.tipo_pedido if mov.pedido else None,  # WEB o MOSTRADOR
                "items":       [],
            }

            if mov.pedido:
                entry["tipo"] = "VENTA"
                for item in mov.pedido.items.select_related("producto").all():
                    nombre = item.producto.nombre if item.producto else item.comentarios
                    entry["items"].append({
                        "cantidad": item.cantidad,
                        "nombre":   nombre,
                        "subtotal": float(item.subtotal),
                    })
                total_ventas += mov.monto
            elif mov.tipo == MovimientoCajaSesion.TIPO_EGRESO:
                entry["tipo"] = "EGRESO"
                total_egresos += mov.monto
            else:
                entry["tipo"] = "INGRESO_MANUAL"

            movimientos.append(entry)

    return JsonResponse({
        "movimientos": movimientos,
        "resumen": {
            "ventas":  float(total_ventas),
            "egresos": float(total_egresos),
            "balance": float(total_ventas - total_egresos),
        }
    })
 
 #────────────────────────────────────────────
 # VISTA AJAX PARA REGISTRAR MOVIMIENTOS MANUALES EN POS
 #────────────────────────────────────────────
@require_POST
def pos_movimiento_manual(request):
    try:
        data   = json.loads(request.body)
        tipo   = data.get("tipo")
        monto  = Decimal(str(data.get("monto", 0)))
        motivo = data.get("motivo", "").strip()

        if not motivo:
            return JsonResponse({"ok": False, "error": "El motivo es obligatorio."})
        if monto <= 0:
            return JsonResponse({"ok": False, "error": "El monto debe ser mayor a 0."})
        if tipo not in [MovimientoCajaSesion.TIPO_INGRESO, MovimientoCajaSesion.TIPO_EGRESO]:
            return JsonResponse({"ok": False, "error": "Tipo inválido."})

        caja_fisica = Caja.objects.filter(activa=True).first()
        sesion      = caja_fisica.sesion_abierta() if caja_fisica else None
        if not sesion:
            return JsonResponse({"ok": False, "error": "No hay sesión de caja abierta."})

        MovimientoCajaSesion.objects.create(
            sesion      = sesion,
            tipo        = tipo,
            monto       = monto,
            descripcion = motivo,
        )
        return JsonResponse({"ok": True})

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})
 
 #────────────────────────────────────────────
 # VISTA AJAX PARA RESUMEN DE CORTE EN POS
 #────────────────────────────────────────────
@require_GET
def pos_corte(request):
    """
    Devuelve el resumen del corte del día:
    - ventas en efectivo (incluye parte efectivo de mixtos)
    - ventas en MP (incluye parte MP de mixtos)
    - ingresos manuales
    - egresos
    - total esperado en caja física (efectivo + ingresos - egresos)
    """
    from django.utils import timezone
    from apps.pedidos.models import Pedido
    from apps.pagos.models import Pago
 
    hoy  = timezone.now().date()
    caja = CajaDiaria.objects.filter(fecha=hoy).first()
 
    ventas_ef    = Decimal("0")
    ventas_mp    = Decimal("0")
    ingresos_man = Decimal("0")
    egresos      = Decimal("0")
 
    if caja:
        for mov in caja.movimientos.select_related("pedido").all():
            if mov.pedido:
                # Desagregar pagos mixtos
                for pago in mov.pedido.pagos.all():
                    if pago.estado == Pago.ESTADO_APROBADO:
                        if pago.tipo == Pago.TIPO_EFECTIVO:
                            ventas_ef += pago.monto
                        else:
                            ventas_mp += pago.monto
            elif mov.tipo == MovimientoCaja.TIPO_INGRESO:
                ingresos_man += mov.monto
            elif mov.tipo == MovimientoCaja.TIPO_EGRESO:
                egresos += mov.monto
 
    efectivo_esperado = ventas_ef + ingresos_man - egresos
    total_esperado    = efectivo_esperado  # MP no entra en caja física
 
    return JsonResponse({
        "ventas_efectivo":    float(ventas_ef),
        "ventas_mp":          float(ventas_mp),
        "ingresos_manuales":  float(ingresos_man),
        "egresos":            float(egresos),
        "efectivo_esperado":  float(efectivo_esperado),
        "total_esperado":     float(total_esperado),
    })


# ─────────────────────────────────────────────
# CAJA SESIÓN (nueva lógica por turnos)
# ─────────────────────────────────────────────

@require_GET
def estado_sesion(request):
    """
    Devuelve la sesión abierta para la caja activa del usuario.
    El frontend consulta esto al cargar el POS para saber si debe
    mostrar el modal de apertura obligatorio.
    """
    caja_id = request.GET.get("caja_id")
    if not caja_id:
        # Sin caja especificada: usar la primera activa
        caja = Caja.objects.filter(activa=True).first()
    else:
        caja = Caja.objects.filter(pk=caja_id, activa=True).first()

    if not caja:
        return JsonResponse({"sesion_abierta": False, "error": "No hay caja configurada."})

    sesion = caja.sesion_abierta()
    if not sesion:
        return JsonResponse({
            "sesion_abierta": False,
            "caja_id":  caja.pk,
            "caja_nombre": caja.nombre,
        })

    return JsonResponse({
        "sesion_abierta": True,
        "sesion_id":      sesion.pk,
        "caja_id":        caja.pk,
        "caja_nombre":    caja.nombre,
        "monto_inicial":  float(sesion.monto_inicial),
        "apertura":       sesion.fecha_apertura.strftime("%d/%m/%Y %H:%M"),
        "usuario":        sesion.usuario_apertura.get_full_name() or sesion.usuario_apertura.username,
    })


@require_POST
def abrir_sesion_caja(request):
    """
    Abre una nueva sesión para la caja.
    Valida que no exista ya una sesión abierta (también lo garantiza
    el UniqueConstraint a nivel BD como segunda línea de defensa).
    """
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "error": "No autenticado."}, status=401)

    try:
        data          = json.loads(request.body)
        caja_id       = data.get("caja_id")
        monto_inicial = Decimal(str(data.get("monto_inicial", "0")))

        if monto_inicial < 0:
            return JsonResponse({"ok": False, "error": "El monto inicial no puede ser negativo."})

        caja = Caja.objects.filter(pk=caja_id, activa=True).first()
        if not caja:
            return JsonResponse({"ok": False, "error": "Caja no encontrada."}, status=404)

        # Primera línea de defensa: check explícito
        if caja.sesion_abierta():
            return JsonResponse({"ok": False, "error": "Ya existe una sesión abierta para esta caja."})

        with transaction.atomic():
            sesion = CajaSesion.objects.create(
                caja             = caja,
                usuario_apertura = request.user,
                monto_inicial    = monto_inicial,
                estado           = CajaSesion.ESTADO_ABIERTA,
            )

        return JsonResponse({
            "ok":        True,
            "sesion_id": sesion.pk,
            "apertura":  sesion.fecha_apertura.strftime("%d/%m/%Y %H:%M"),
        })

    except IntegrityError:
        # Segunda línea de defensa: UniqueConstraint de BD
        return JsonResponse({"ok": False, "error": "Ya existe una sesión abierta para esta caja."})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@require_GET
def datos_corte_sesion(request):
    """
    Devuelve el resumen de la sesión activa para mostrar en el modal de cierre.
    El frontend NUNCA calcula efectivo_esperado; siempre viene del backend.
    """
    sesion_id = request.GET.get("sesion_id")
    if not sesion_id:
        # Intentar obtener la sesión activa de la primera caja
        caja   = Caja.objects.filter(activa=True).first()
        sesion = caja.sesion_abierta() if caja else None
    else:
        sesion = CajaSesion.objects.filter(pk=sesion_id, estado=CajaSesion.ESTADO_ABIERTA).first()

    if not sesion:
        return JsonResponse({"ok": False, "error": "No hay sesión abierta."}, status=404)

    datos = sesion.datos_corte()
    datos["ok"]       = True
    datos["sesion_id"]= sesion.pk
    return JsonResponse(datos)


@require_POST
def cerrar_sesion_caja(request):
    """
    Cierra la sesión de caja.
    - Recalcula efectivo_esperado en backend (nunca confía en el frontend).
    - Usa transacción para evitar race conditions.
    - Bloquea doble submit con select_for_update.
    """
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "error": "No autenticado."}, status=401)

    try:
        data          = json.loads(request.body)
        sesion_id     = data.get("sesion_id")
        efectivo_real = data.get("efectivo_real")

        if efectivo_real is None:
            return JsonResponse({"ok": False, "error": "El efectivo real es obligatorio."})

        try:
            efectivo_real = Decimal(str(efectivo_real))
        except Exception:
            return JsonResponse({"ok": False, "error": "Monto inválido."})

        if efectivo_real < 0:
            return JsonResponse({"ok": False, "error": "El monto no puede ser negativo."})

        with transaction.atomic():
            # select_for_update evita race condition si dos requests llegan simultáneos
            sesion = (
                CajaSesion.objects
                .select_for_update()
                .filter(pk=sesion_id)
                .first()
            )
            if not sesion:
                return JsonResponse({"ok": False, "error": "Sesión no encontrada."}, status=404)

            if sesion.estado == CajaSesion.ESTADO_CERRADA:
                return JsonResponse({"ok": False, "error": "La sesión ya fue cerrada."})

            sesion.cerrar(efectivo_real, request.user)

        return JsonResponse({
            "ok":               True,
            "efectivo_esperado":float(sesion.efectivo_esperado),
            "efectivo_real":    float(sesion.efectivo_real),
            "diferencia":       float(sesion.diferencia),
        })

    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({"ok": False, "error": str(e)}, status=500)